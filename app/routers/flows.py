from fastapi import APIRouter, Depends, HTTPException, status, Request
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func
from typing import Optional, List, Dict
from datetime import date, datetime, timedelta

from app.database import get_db
from app import models
from app.dependencies import get_current_active_user

router = APIRouter(prefix="/flows", tags=["flows"])
templates = Jinja2Templates(directory="app/templates")


@router.get("/", response_class=HTMLResponse)
async def flows_page(
    request: Request,
    db: Session = Depends(get_db)
):
    """Страница со списком потоков (папок)"""
    current_user = await get_current_active_user(request, db)
    
    # Получаем все письма, доступные пользователю
    if current_user.is_admin:
        correspondences = db.query(models.Correspondence).all()
    else:
        correspondences = db.query(models.Correspondence).filter(
            or_(
                models.Correspondence.executor_id == current_user.id,
                models.Correspondence.created_by_id == current_user.id,
                models.Correspondence.id.in_(
                    db.query(models.CorrespondenceTransfer.correspondence_id).filter(
                        models.CorrespondenceTransfer.transferred_to_id == current_user.id,
                        models.CorrespondenceTransfer.is_active == True
                    )
                )
            )
        ).all()
    
    # Группируем по потокам в Python (проще и надежнее)
    flows_dict = {}
    for corr in correspondences:
        flow_name = corr.flow or "без потока"
        if flow_name not in flows_dict:
            flows_dict[flow_name] = {"total": 0, "completed": 0}
        flows_dict[flow_name]["total"] += 1
        if corr.status == models.CorrespondenceStatus.COMPLETED:
            flows_dict[flow_name]["completed"] += 1
    
    # Преобразуем в список словарей
    flows = []
    for flow_name, data in flows_dict.items():
        total = data["total"]
        completed = data["completed"]
        in_progress = total - completed
        
        flows.append({
            "name": flow_name,
            "total": total,
            "completed": completed,
            "in_progress": in_progress,
            "progress_percent": round((completed / total * 100) if total > 0 else 0)
        })
    
    # Сортируем по имени потока
    flows.sort(key=lambda x: x["name"])
    
    return templates.TemplateResponse(
        "flows.html",
        {
            "request": request,
            "current_user": current_user,
            "flows": flows,
            "today": date.today()
        }
    )


@router.get("/{flow_name}", response_class=HTMLResponse)
async def flow_detail(
    request: Request,
    flow_name: str,
    db: Session = Depends(get_db)
):
    """Страница с письмами конкретного потока"""
    current_user = await get_current_active_user(request, db)
    
    # Получаем письма потока
    if current_user.is_admin:
        correspondences = db.query(models.Correspondence).filter(
            models.Correspondence.flow == flow_name
        ).order_by(models.Correspondence.created_at.desc()).all()
    else:
        correspondences = db.query(models.Correspondence).filter(
            and_(
                models.Correspondence.flow == flow_name,
                or_(
                    models.Correspondence.executor_id == current_user.id,
                    models.Correspondence.created_by_id == current_user.id,
                    models.Correspondence.id.in_(
                        db.query(models.CorrespondenceTransfer.correspondence_id).filter(
                            models.CorrespondenceTransfer.transferred_to_id == current_user.id,
                            models.CorrespondenceTransfer.is_active == True
                        )
                    )
                )
            )
        ).order_by(models.Correspondence.created_at.desc()).all()
    
    # Статистика по потоку
    total = len(correspondences)
    completed = sum(1 for c in correspondences if c.status == models.CorrespondenceStatus.COMPLETED)
    in_progress = total - completed
    progress_percent = round((completed / total * 100) if total > 0 else 0)
    
    return templates.TemplateResponse(
        "flow_detail.html",
        {
            "request": request,
            "current_user": current_user,
            "flow_name": flow_name,
            "correspondences": correspondences,
            "total": total,
            "completed": completed,
            "in_progress": in_progress,
            "progress_percent": progress_percent,
            "today": date.today(),
            "statuses": [s.value for s in models.CorrespondenceStatus]
        }
    )


@router.get("/{flow_name}/export")
async def export_flow_to_excel(
    request: Request,
    flow_name: str,
    db: Session = Depends(get_db)
):
    """Экспорт писем конкретного потока в Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    current_user = await get_current_active_user(request, db)
    
    # Получаем письма потока
    if current_user.is_admin:
        correspondences = db.query(models.Correspondence).filter(
            models.Correspondence.flow == flow_name
        ).order_by(models.Correspondence.created_at.desc()).all()
    else:
        correspondences = db.query(models.Correspondence).filter(
            and_(
                models.Correspondence.flow == flow_name,
                or_(
                    models.Correspondence.executor_id == current_user.id,
                    models.Correspondence.created_by_id == current_user.id,
                    models.Correspondence.id.in_(
                        db.query(models.CorrespondenceTransfer.correspondence_id).filter(
                            models.CorrespondenceTransfer.transferred_to_id == current_user.id,
                            models.CorrespondenceTransfer.is_active == True
                        )
                    )
                )
            )
        ).order_by(models.Correspondence.created_at.desc()).all()
    
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = f"Поток {flow_name}"
    
    # Стили
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_side = Side(style='thin', color='CCCCCC')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Заголовки
    headers = [
        '№', 'Входящий номер', 'Дата входящего', 'Дата поступления',
        'Отправитель', 'Содержание', 'Исполнитель', 'Срок исполнения',
        'Статус', 'Отправлено', 'Кому', 'Контроль', 'Отчет', 'Дата отчета'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Заполняем данными
    for row_idx, corr in enumerate(correspondences, 2):
        executor_name = "Не назначен"
        if corr.executor:
            executor_name = corr.executor.full_name
        elif corr.created_by:
            executor_name = corr.created_by.full_name
        
        data = [
            row_idx - 1,
            corr.incoming_number or '',
            corr.incoming_date.strftime('%d.%m.%Y') if corr.incoming_date else '',
            corr.received_date.strftime('%d.%m.%Y') if corr.received_date else '',
            corr.sender or '',
            corr.content or '',
            executor_name,
            corr.deadline.strftime('%d.%m.%Y') if corr.deadline else '',
            corr.status.value if corr.status else '',
            corr.sent_info or '',
            corr.to_whom or '',
            corr.control or '',
            corr.report or '',
            corr.report_date.strftime('%d.%m.%Y') if corr.report_date else ''
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = border
    
    # Автоматическая ширина колонок
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    safe_flow_name = flow_name.replace('/', '_').replace('\\', '_')
    filename = f"flow_{safe_flow_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )