from fastapi import APIRouter, Depends, HTTPException, status, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from typing import Optional
from datetime import date, datetime
from io import BytesIO

from app.database import get_db
from app import models
from app.dependencies import get_current_active_user

router = APIRouter(prefix="/requests", tags=["requests"])
templates = Jinja2Templates(directory="app/templates")


@router.get("/", response_class=HTMLResponse)
async def requests_page(
    request: Request,
    request_type_filter: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Страница со списком писем-запросов"""
    current_user = await get_current_active_user(request, db)
    
    # Базовый запрос - только письма с заполненным request_type
    if current_user.is_admin:
        query = db.query(models.Correspondence).filter(
            models.Correspondence.request_type.isnot(None)
        )
    else:
        query = db.query(models.Correspondence).filter(
            and_(
                models.Correspondence.request_type.isnot(None),
                or_(
                    models.Correspondence.executor_id == current_user.id,
                    models.Correspondence.created_by_id == current_user.id,
                    models.Correspondence.id.in_(
                        db.query(models.CorrespondenceTransfer.correspondence_id).filter(
                            models.CorrespondenceTransfer.transferred_to_id == current_user.id,
                            models.CorrespondenceTransfer.is_active == True
                        )
                    )
                )
            )
        )
    
    # Фильтр по типу запроса
    if request_type_filter:
        query = query.filter(models.Correspondence.request_type == request_type_filter)
    
    correspondences = query.order_by(models.Correspondence.created_at.desc()).all()
    
    # Статистика
    total = len(correspondences)
    own_purposes = sum(1 for c in correspondences if c.request_type == models.RequestType.OWN_PURPOSES)
    ministry = sum(1 for c in correspondences if c.request_type == models.RequestType.MINISTRY_OF_EDUCATION)
    
    # Типы запросов для фильтра
    request_types = [
        {"value": models.RequestType.OWN_PURPOSES.value, "label": "Для собственных целей"},
        {"value": models.RequestType.MINISTRY_OF_EDUCATION.value, "label": "Для Министерства просвещения России"}
    ]
    
    return templates.TemplateResponse(
        "requests.html",
        {
            "request": request,
            "current_user": current_user,
            "correspondences": correspondences,
            "total": total,
            "own_purposes": own_purposes,
            "ministry": ministry,
            "request_types": request_types,
            "request_type_filter": request_type_filter,
            "today": date.today(),
            "statuses": [s.value for s in models.CorrespondenceStatus]
        }
    )


@router.get("/export")
async def export_requests_to_excel(
    request: Request,
    request_type_filter: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Экспорт писем-запросов в Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    current_user = await get_current_active_user(request, db)
    
    # Базовый запрос - только письма с заполненным request_type
    if current_user.is_admin:
        query = db.query(models.Correspondence).filter(
            models.Correspondence.request_type.isnot(None)
        )
    else:
        query = db.query(models.Correspondence).filter(
            and_(
                models.Correspondence.request_type.isnot(None),
                or_(
                    models.Correspondence.executor_id == current_user.id,
                    models.Correspondence.created_by_id == current_user.id,
                    models.Correspondence.id.in_(
                        db.query(models.CorrespondenceTransfer.correspondence_id).filter(
                            models.CorrespondenceTransfer.transferred_to_id == current_user.id,
                            models.CorrespondenceTransfer.is_active == True
                        )
                    )
                )
            )
        )
    
    if request_type_filter:
        query = query.filter(models.Correspondence.request_type == request_type_filter)
    
    correspondences = query.order_by(models.Correspondence.created_at.desc()).all()
    
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Письма-запросы"
    
    # Стили
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_side = Side(style='thin', color='CCCCCC')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Заголовки
    headers = [
        '№', 'Входящий номер', 'Дата входящего', 'Дата поступления',
        'Отправитель', 'Содержание', 'Исполнитель', 'Срок исполнения',
        'Статус', 'Тип запроса', 'Примечание', 'Отправлено', 'Кому', 'Контроль'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Заполняем данными
    for row_idx, corr in enumerate(correspondences, 2):
        executor_name = "Не назначен"
        if corr.executor:
            executor_name = corr.executor.full_name
        elif corr.created_by:
            executor_name = corr.created_by.full_name
        
        request_type_value = ""
        if corr.request_type == models.RequestType.OWN_PURPOSES:
            request_type_value = "Для собственных целей"
        elif corr.request_type == models.RequestType.MINISTRY_OF_EDUCATION:
            request_type_value = "Для Министерства просвещения России"
        
        data = [
            row_idx - 1,
            corr.incoming_number or '',
            corr.incoming_date.strftime('%d.%m.%Y') if corr.incoming_date else '',
            corr.received_date.strftime('%d.%m.%Y') if corr.received_date else '',
            corr.sender or '',
            corr.content or '',
            executor_name,
            corr.deadline.strftime('%d.%m.%Y') if corr.deadline else '',
            corr.status.value if corr.status else '',
            request_type_value,
            corr.request_note or '',
            corr.sent_info or '',
            corr.to_whom or '',
            corr.control or ''
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = border
    
    # Автоматическая ширина колонок
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"requests_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )