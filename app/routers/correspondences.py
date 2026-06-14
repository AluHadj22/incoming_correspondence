from fastapi import APIRouter, Depends, HTTPException, status, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func
from typing import Optional, List
from datetime import date, datetime, timedelta
import re

from app.database import get_db
from app import models, schemas, auth
from app.dependencies import (
    get_current_active_user, get_correspondence_or_404, 
    check_correspondence_access, get_user_correspondences,
    get_dashboard_stats, get_department_users
)

router = APIRouter(prefix="/api/correspondences", tags=["correspondences"])
templates = Jinja2Templates(directory="app/templates")


# --- Функция для извлечения потока из входящего номера ---
def extract_flow_from_number(incoming_number: str) -> str:
    """
    Извлекает поток из входящего номера.
    Берет ВСЕ символы после первого слеша "/"
    Примеры:
    "642/07-25" -> "07-25"
    "123/ABC-123" -> "ABC-123"
    "456/ДСП-01" -> "ДСП-01"
    "789/Письмо_от_ФСБ" -> "Письмо_от_ФСБ"
    "111/22.11.2025" -> "22.11.2025"
    """
    if not incoming_number:
        return "без потока"
    
    # Ищем всё, что идет после первого слеша "/"
    if '/' in incoming_number:
        parts = incoming_number.split('/', 1)
        if len(parts) > 1 and parts[1].strip():
            return parts[1].strip()
    
    return "без потока"


# --- HTML страницы ---
@router.get("/", response_class=HTMLResponse)
async def correspondences_page(
    request: Request,
    status_filter: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Страница со списком писем"""
    current_user = await get_current_active_user(request, db)
    correspondences = await get_user_correspondences(request, db, status_filter)
    
    # Получаем список пользователей для фильтра (только из отдела текущего пользователя)
    users = get_department_users(current_user.department, db)
    
    return templates.TemplateResponse(
        "correspondences.html",
        {
            "request": request,
            "correspondences": correspondences,
            "current_user": current_user,
            "status_filter": status_filter,
            "users": users,
            "statuses": [s.value for s in models.CorrespondenceStatus],
            "today": date.today()
        }
    )


@router.get("/create", response_class=HTMLResponse)
async def create_correspondence_form(
    request: Request,
    db: Session = Depends(get_db)
):
    """Форма создания нового письма"""
    current_user = await get_current_active_user(request, db)
    # Если админ - показывает всех пользователей, иначе только из своего отдела
    if current_user.is_admin:
        users = db.query(models.User).filter(models.User.is_active == True).all()
    else:
        users = get_department_users(current_user.department, db)
    
    # Типы запросов для выпадающего списка
    request_types = [
        {"value": models.RequestType.OWN_PURPOSES.value, "label": "Для собственных целей"},
        {"value": models.RequestType.MINISTRY_OF_EDUCATION.value, "label": "Для Министерства просвещения России"}
    ]
    
    return templates.TemplateResponse(
        "correspondence_form.html",
        {
            "request": request,
            "current_user": current_user,
            "users": users,
            "statuses": [s.value for s in models.CorrespondenceStatus],
            "request_types": request_types,
            "is_edit": False
        }
    )


# ВАЖНО: эндпоинт export должен быть ПЕРЕД маршрутами с {correspondence_id}
@router.get("/export")
async def export_correspondences_to_excel(
    request: Request,
    status_filter: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Экспорт писем в Excel файл"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    current_user = await get_current_active_user(request, db)
    
    # Получаем письма с фильтром
    correspondences = await get_user_correspondences(request, db, status_filter)
    
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Входящая корреспонденция"
    
    # Стили
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_side = Side(style='thin', color='CCCCCC')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Заголовки (добавили колонки для запросов)
    headers = [
        '№', 'Входящий номер', 'Поток', 'Дата входящего', 'Дата поступления',
        'Отправитель', 'Содержание', 'Исполнитель', 'Срок исполнения',
        'Статус', 'Тип запроса', 'Примечание к запросу',
        'Отправлено', 'Кому', 'Контроль', 'Отчет', 'Дата отчета'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Заполняем данными
    for row_idx, corr in enumerate(correspondences, 2):
        # Получаем имя исполнителя
        executor_name = "Не назначен"
        if corr.executor:
            executor_name = corr.executor.full_name
        elif corr.created_by:
            executor_name = corr.created_by.full_name
        
        # Получаем тип запроса
        request_type_value = ""
        if corr.request_type == models.RequestType.OWN_PURPOSES:
            request_type_value = "Для собственных целей"
        elif corr.request_type == models.RequestType.MINISTRY_OF_EDUCATION:
            request_type_value = "Для Министерства просвещения России"
        
        data = [
            row_idx - 1,  # №
            corr.incoming_number or '',
            corr.flow or extract_flow_from_number(corr.incoming_number),
            corr.incoming_date.strftime('%d.%m.%Y') if corr.incoming_date else '',
            corr.received_date.strftime('%d.%m.%Y') if corr.received_date else '',
            corr.sender or '',
            corr.content or '',
            executor_name,
            corr.deadline.strftime('%d.%m.%Y') if corr.deadline else '',
            corr.status.value if corr.status else '',
            request_type_value,
            corr.request_note or '',
            corr.sent_info or '',
            corr.to_whom or '',
            corr.control or '',
            corr.report or '',
            corr.report_date.strftime('%d.%m.%Y') if corr.report_date else ''
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = border
    
    # Автоматическая ширина колонок
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Заморозка первой строки
    ws.freeze_panes = 'A2'
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Формируем имя файла с текущей датой
    filename = f"correspondences_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{correspondence_id}/edit", response_class=HTMLResponse)
async def edit_correspondence_form(
    request: Request,
    correspondence_id: int,
    db: Session = Depends(get_db)
):
    """Форма редактирования письма"""
    current_user = await get_current_active_user(request, db)
    correspondence = await check_correspondence_access(request, correspondence_id, db)
    # Если админ - показывает всех пользователей, иначе только из своего отдела
    if current_user.is_admin:
        users = db.query(models.User).filter(models.User.is_active == True).all()
    else:
        users = get_department_users(current_user.department, db)
    
    # Типы запросов для выпадающего списка
    request_types = [
        {"value": models.RequestType.OWN_PURPOSES.value, "label": "Для собственных целей"},
        {"value": models.RequestType.MINISTRY_OF_EDUCATION.value, "label": "Для Министерства просвещения России"}
    ]
    
    return templates.TemplateResponse(
        "correspondence_form.html",
        {
            "request": request,
            "current_user": current_user,
            "users": users,
            "correspondence": correspondence,
            "statuses": [s.value for s in models.CorrespondenceStatus],
            "request_types": request_types,
            "is_edit": True
        }
    )


@router.get("/{correspondence_id}", response_class=HTMLResponse)
async def view_correspondence(
    request: Request,
    correspondence_id: int,
    db: Session = Depends(get_db)
):
    """Просмотр деталей письма"""
    current_user = await get_current_active_user(request, db)
    correspondence = await check_correspondence_access(request, correspondence_id, db)
    
    # Получаем историю передач
    transfers = db.query(models.CorrespondenceTransfer).filter(
        models.CorrespondenceTransfer.correspondence_id == correspondence_id,
        models.CorrespondenceTransfer.is_active == True
    ).all()
    
    # Получаем доступных пользователей для передачи
    # Если админ - показывает всех пользователей, иначе только из своего отдела
    if current_user.is_admin:
        available_users = db.query(models.User).filter(models.User.is_active == True).all()
    else:
        available_users = get_department_users(current_user.department, db)
    
    return templates.TemplateResponse(
        "correspondence_detail.html",
        {
            "request": request,
            "correspondence": correspondence,
            "current_user": current_user,
            "transfers": transfers,
            "available_users": available_users,
            "statuses": [s.value for s in models.CorrespondenceStatus],
            "today": date.today()
        }
    )


# --- API эндпоинты (CRUD) ---
@router.post("/create", response_class=HTMLResponse)
async def create_correspondence(
    request: Request,
    incoming_number: str = Form(...),
    incoming_date: str = Form(...),
    received_date: str = Form(...),
    sender: str = Form(...),
    content: str = Form(...),
    deadline: str = Form(...),
    status_value: str = Form(..., alias="status"),
    executor_id: Optional[int] = Form(None),
    sent_info: Optional[str] = Form(None),
    to_whom: Optional[str] = Form(None),
    control: Optional[str] = Form(None),
    report: Optional[str] = Form(None),
    report_date: Optional[str] = Form(None),
    request_type: Optional[str] = Form(None),
    request_note: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Создание нового письма"""
    current_user = await get_current_active_user(request, db)
    
    try:
        # Преобразуем строки в даты
        incoming_date_obj = datetime.strptime(incoming_date, "%Y-%m-%d").date()
        received_date_obj = datetime.strptime(received_date, "%Y-%m-%d").date()
        deadline_obj = datetime.strptime(deadline, "%Y-%m-%d").date()
        report_date_obj = datetime.strptime(report_date, "%Y-%m-%d").date() if report_date else None
        
        # Преобразуем статус в enum
        status_enum = None
        for s in models.CorrespondenceStatus:
            if s.value == status_value:
                status_enum = s
                break
        
        # Преобразуем тип запроса в enum
        request_type_enum = None
        if request_type:
            for rt in models.RequestType:
                if rt.value == request_type:
                    request_type_enum = rt
                    break
        
        # Извлекаем поток из входящего номера
        flow = extract_flow_from_number(incoming_number)
        
        new_correspondence = models.Correspondence(
            incoming_number=incoming_number,
            incoming_date=incoming_date_obj,
            received_date=received_date_obj,
            sender=sender,
            content=content,
            deadline=deadline_obj,
            status=status_enum or models.CorrespondenceStatus.PENDING,
            executor_id=executor_id,
            created_by_id=current_user.id,
            sent_info=sent_info,
            to_whom=to_whom,
            control=control,
            report=report,
            report_date=report_date_obj,
            flow=flow,
            request_type=request_type_enum,
            request_note=request_note if request_note else None
        )
        
        db.add(new_correspondence)
        db.commit()
        db.refresh(new_correspondence)
        
        return RedirectResponse(
            url=f"/api/correspondences/{new_correspondence.id}",
            status_code=303
        )
        
    except Exception as e:
        if current_user.is_admin:
            users = db.query(models.User).filter(models.User.is_active == True).all()
        else:
            users = get_department_users(current_user.department, db)
        
        request_types = [
            {"value": models.RequestType.OWN_PURPOSES.value, "label": "Для собственных целей"},
            {"value": models.RequestType.MINISTRY_OF_EDUCATION.value, "label": "Для Министерства просвещения России"}
        ]
        
        return templates.TemplateResponse(
            "correspondence_form.html",
            {
                "request": request,
                "current_user": current_user,
                "users": users,
                "statuses": [s.value for s in models.CorrespondenceStatus],
                "request_types": request_types,
                "is_edit": False,
                "error": f"Ошибка при создании: {str(e)}"
            },
            status_code=400
        )


@router.post("/{correspondence_id}/edit")
async def update_correspondence(
    request: Request,
    correspondence_id: int,
    incoming_number: str = Form(...),
    incoming_date: str = Form(...),
    received_date: str = Form(...),
    sender: str = Form(...),
    content: str = Form(...),
    deadline: str = Form(...),
    status_value: str = Form(..., alias="status"),
    executor_id: Optional[int] = Form(None),
    sent_info: Optional[str] = Form(None),
    to_whom: Optional[str] = Form(None),
    control: Optional[str] = Form(None),
    report: Optional[str] = Form(None),
    report_date: Optional[str] = Form(None),
    request_type: Optional[str] = Form(None),
    request_note: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Обновление письма"""
    current_user = await get_current_active_user(request, db)
    correspondence = await check_correspondence_access(request, correspondence_id, db)
    
    try:
        # Преобразуем даты
        incoming_date_obj = datetime.strptime(incoming_date, "%Y-%m-%d").date()
        received_date_obj = datetime.strptime(received_date, "%Y-%m-%d").date()
        deadline_obj = datetime.strptime(deadline, "%Y-%m-%d").date()
        report_date_obj = datetime.strptime(report_date, "%Y-%m-%d").date() if report_date else None
        
        # Обновляем статус
        status_enum = None
        for s in models.CorrespondenceStatus:
            if s.value == status_value:
                status_enum = s
                break
        
        # Преобразуем тип запроса в enum
        request_type_enum = None
        if request_type:
            for rt in models.RequestType:
                if rt.value == request_type:
                    request_type_enum = rt
                    break
        
        # Если статус меняется на "Исполнено", устанавливаем дату выполнения
        completed_at = correspondence.completed_at
        if status_enum == models.CorrespondenceStatus.COMPLETED and correspondence.status != models.CorrespondenceStatus.COMPLETED:
            completed_at = datetime.utcnow()
        
        # Обновляем поток при изменении номера
        flow = extract_flow_from_number(incoming_number)
        
        # Обновляем поля
        correspondence.incoming_number = incoming_number
        correspondence.incoming_date = incoming_date_obj
        correspondence.received_date = received_date_obj
        correspondence.sender = sender
        correspondence.content = content
        correspondence.deadline = deadline_obj
        correspondence.status = status_enum or models.CorrespondenceStatus.PENDING
        correspondence.executor_id = executor_id
        correspondence.sent_info = sent_info
        correspondence.to_whom = to_whom
        correspondence.control = control
        correspondence.report = report
        correspondence.report_date = report_date_obj
        correspondence.completed_at = completed_at
        correspondence.updated_at = datetime.utcnow()
        correspondence.flow = flow
        correspondence.request_type = request_type_enum
        correspondence.request_note = request_note if request_note else None
        
        db.commit()
        db.refresh(correspondence)
        
        return RedirectResponse(
            url=f"/api/correspondences/{correspondence.id}",
            status_code=303
        )
        
    except Exception as e:
        if current_user.is_admin:
            users = db.query(models.User).filter(models.User.is_active == True).all()
        else:
            users = get_department_users(current_user.department, db)
        
        request_types = [
            {"value": models.RequestType.OWN_PURPOSES.value, "label": "Для собственных целей"},
            {"value": models.RequestType.MINISTRY_OF_EDUCATION.value, "label": "Для Министерства просвещения России"}
        ]
        
        return templates.TemplateResponse(
            "correspondence_form.html",
            {
                "request": request,
                "current_user": current_user,
                "users": users,
                "correspondence": correspondence,
                "statuses": [s.value for s in models.CorrespondenceStatus],
                "request_types": request_types,
                "is_edit": True,
                "error": f"Ошибка при обновлении: {str(e)}"
            },
            status_code=400
        )


@router.post("/{correspondence_id}/transfer")
async def transfer_correspondence(
    request: Request,
    correspondence_id: int,
    transferred_to_id: int = Form(...),
    note: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Передача письма другому сотруднику"""
    current_user = await get_current_active_user(request, db)
    correspondence = await check_correspondence_access(request, correspondence_id, db)
    
    # Проверяем, существует ли пользователь
    target_user = db.query(models.User).filter(models.User.id == transferred_to_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Деактивируем предыдущие активные передачи для этого письма
    active_transfers = db.query(models.CorrespondenceTransfer).filter(
        and_(
            models.CorrespondenceTransfer.correspondence_id == correspondence_id,
            models.CorrespondenceTransfer.is_active == True
        )
    ).all()
    
    for transfer in active_transfers:
        transfer.is_active = False
    
    # Создаем новую передачу
    new_transfer = models.CorrespondenceTransfer(
        correspondence_id=correspondence_id,
        transferred_by_id=current_user.id,
        transferred_to_id=transferred_to_id,
        note=note,
        is_active=True
    )
    
    # Обновляем статус письма
    correspondence.status = models.CorrespondenceStatus.TRANSFERRED
    # Автоматически назначаем исполнителя - того, кому передали письмо
    correspondence.executor_id = transferred_to_id
    
    db.add(new_transfer)
    db.commit()
    
    return RedirectResponse(
        url=f"/api/correspondences/{correspondence_id}",
        status_code=303
    )


@router.post("/{correspondence_id}/delete")
async def delete_correspondence(
    request: Request,
    correspondence_id: int,
    db: Session = Depends(get_db)
):
    """Удаление письма (только для админа или создателя)"""
    current_user = await get_current_active_user(request, db)
    correspondence = await check_correspondence_access(request, correspondence_id, db)
    
    if not current_user.is_admin and correspondence.created_by_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin or creator can delete correspondence"
        )
    
    # Удаляем связанные передачи
    db.query(models.CorrespondenceTransfer).filter(
        models.CorrespondenceTransfer.correspondence_id == correspondence_id
    ).delete()
    
    # Удаляем письмо
    db.delete(correspondence)
    db.commit()
    
    return RedirectResponse(
        url="/api/correspondences/",
        status_code=303
    )